"""Document text extraction and upload catalog helpers."""

import io
import os
import re
import tempfile
from email import policy
from email.message import EmailMessage, Message
from email.parser import BytesParser
from typing import Callable


class TextExtractionService:
    """Extract text from supported document types and describe uploaded files."""

    MIME_TYPE_FUNC_MAPPING: dict[str, str] = {
        "text/plain": "read_text",
        "text/html": "read_html",
        "application/pdf": "read_pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "read_docx",
        "application/vnd.ms-excel": "read_excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "read_excel",
        "application/vnd.ms-outlook": "read_outlook",
        "application/octet-stream": "read_outlook",
        "message/rfc822": "read_rfc822",
        "image/jpeg": "read_image",
        "image/jpg": "read_image",
        "image/png": "read_image",
        "image/gif": "read_image",
        "image/bmp": "read_image",
        "image/tiff": "read_image",
        "image/webp": "read_image",
    }
    EXTENSION_TO_MIME: dict[str, str] = {
        ".txt": "text/plain",
        ".html": "text/html",
        ".htm": "text/html",
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".msg": "application/vnd.ms-outlook",
        ".eml": "message/rfc822",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
        ".webp": "image/webp",
    }
    CANONICAL_MIME_TO_EXTENSION: dict[str, str] = {
        "text/plain": ".txt",
        "text/html": ".html",
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.ms-excel": ".xls",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/vnd.ms-outlook": ".msg",
        "message/rfc822": ".eml",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/tiff": ".tiff",
        "image/webp": ".webp",
    }
    EXTENSION_ALIASES: dict[str, str] = {
        ".htm": ".html",
        ".jpeg": ".jpg",
        ".tif": ".tiff",
    }
    _tiktoken_encoding = None

    def __init__(self) -> None:
        """Initialize extractor functions and supported MIME metadata."""
        self.extractors: dict[str, Callable[..., tuple[str, int]]] = {
            "read_text": self.read_text,
            "read_html": self.read_html,
            "read_pdf": self.read_pdf,
            "read_docx": self.read_docx,
            "read_excel": self.read_excel,
            "read_outlook": self.read_outlook,
            "read_rfc822": self.read_rfc822,
            "read_image": self.read_image,
        }
        self.ext_to_mime: dict[str, str] = dict(self.EXTENSION_TO_MIME)
        self.mime_type_func_mapping: dict[str, str] = dict(self.MIME_TYPE_FUNC_MAPPING)
        self.extension_to_extractor: dict[str, str] = self._build_extension_to_extractor()

    def _build_extension_to_extractor(self) -> dict[str, str]:
        """Build the extension-to-extractor lookup from MIME declarations."""
        mapping: dict[str, str] = {}
        for extension, mime_type in self.ext_to_mime.items():
            extractor_name = self._mime_type_to_extractor_name(
                mime_type, file_name=f"file{extension}"
            )
            if extractor_name:
                mapping[extension] = extractor_name
        return mapping

    def normalize_extension(self, extension: str) -> str:
        """Normalize an extension into the canonical lookup form.

        Args:
            extension: Extension with or without a leading dot.

        Returns:
            Canonical lowercase extension including the leading dot.
        """
        if not extension:
            return ""
        normalized = extension.lower()
        if not normalized.startswith("."):
            normalized = f".{normalized}"
        return self.EXTENSION_ALIASES.get(normalized, normalized)

    def get_extension(self, file_name: str) -> str:
        """Return the normalized extension for a file name."""
        return self.normalize_extension(os.path.splitext(file_name)[1].lower())

    def _normalize_mime_type(self, mime_type: str | None) -> str | None:
        """Normalize a MIME type string for case-insensitive lookups."""
        if not mime_type:
            return None
        return mime_type.lower().strip()

    def _mime_type_to_extractor_name(
        self,
        mime_type: str | None,
        file_name: str = "",
    ) -> str | None:
        """Resolve a MIME type to an extractor name.

        `application/octet-stream` is only accepted when the file extension
        indicates an Outlook `.msg` payload, which is common for browsers that
        upload `.msg` files without a more specific MIME type.
        """
        normalized_mime = self._normalize_mime_type(mime_type)
        if not normalized_mime:
            return None
        if normalized_mime == "application/octet-stream":
            return (
                self.mime_type_func_mapping[normalized_mime]
                if self.get_extension(file_name) == ".msg"
                else None
            )
        return self.mime_type_func_mapping.get(normalized_mime)

    def resolve_extension(self, file_name: str, mime_type: str | None = None) -> str:
        """Resolve the best supported extension from file name and MIME type."""
        extension = self.get_extension(file_name)
        if extension in self.extension_to_extractor:
            return extension

        normalized_mime = self._normalize_mime_type(mime_type)
        if normalized_mime:
            return self.CANONICAL_MIME_TO_EXTENSION.get(normalized_mime, extension)

        return extension

    def resolve_mime_type(self, file_name: str, mime_type: str | None = None) -> str:
        """Resolve the canonical MIME type for a file."""
        extension = self.resolve_extension(file_name, mime_type)
        if extension in self.ext_to_mime:
            return self.ext_to_mime[extension]

        normalized_mime = self._normalize_mime_type(mime_type)
        if normalized_mime and self._mime_type_to_extractor_name(normalized_mime, file_name):
            return normalized_mime

        return "application/octet-stream"

    def resolve_extractor_name(
        self,
        type_identifier: str | None = None,
        *,
        file_name: str = "",
        mime_type: str | None = None,
    ) -> str | None:
        """Resolve an extractor name from an extension or MIME type."""
        normalized_identifier = (type_identifier or "").lower().strip()
        if normalized_identifier.startswith("."):
            return self.extension_to_extractor.get(self.normalize_extension(normalized_identifier))

        extractor_from_identifier = self._mime_type_to_extractor_name(
            normalized_identifier or None,
            file_name=file_name,
        )
        if extractor_from_identifier:
            return extractor_from_identifier

        extension = self.resolve_extension(file_name, mime_type)
        extractor_from_extension = self.extension_to_extractor.get(extension)
        if extractor_from_extension:
            return extractor_from_extension

        return self._mime_type_to_extractor_name(mime_type, file_name=file_name)

    def is_supported_upload(self, file_name: str, mime_type: str | None = None) -> bool:
        """Return whether the file can be extracted with the current registry.

        Args:
            file_name: Uploaded or staged file name.
            mime_type: Optional MIME type from the client/upload metadata.

        Returns:
            `True` when an extractor is available for the file.
        """
        return self.resolve_extractor_name(file_name=file_name, mime_type=mime_type) is not None

    @property
    def allowed_extensions(self) -> set[str]:
        """Return all supported file extensions."""
        return set(self.extension_to_extractor.keys())

    def _decode_text_bytes(self, file_bytes: bytes) -> str:
        """Decode plain-text bytes with a few common fallbacks."""
        for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return file_bytes.decode("utf-8", errors="replace")

    def _normalize_text(self, text: str) -> str:
        """Collapse overly noisy whitespace while preserving paragraphs."""
        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        normalized = "\n".join(line.rstrip() for line in normalized.split("\n"))
        normalized = re.sub(r"\n{3,}", "\n\n", normalized)
        return normalized.strip()

    def _html_to_text(self, html_text: str) -> str:
        """Convert HTML markup into readable plain text."""
        from bs4 import BeautifulSoup

        soup = BeautifulSoup(html_text, "html.parser")
        for tag_name in ("script", "style", "noscript"):
            for element in soup.find_all(tag_name):
                element.decompose()
        return self._normalize_text(soup.get_text(separator="\n"))

    def _stringify_row_cells(self, values: list[object]) -> str:
        """Format a workbook row into a tab-delimited string."""
        cells = [str(value) if value is not None else "" for value in values]
        return "\t".join(cells)

    def extract_segments(
        self,
        type_identifier: str,
        file_bytes: bytes,
        *,
        file_name: str = "",
        mime_type: str | None = None,
    ) -> tuple[list[str], int]:
        """Extract ordered text segments for a supported file type.

        This helper keeps page-, sheet-, and message-aware extraction
        centralized so other services do not need file-type-specific logic.

        Args:
            type_identifier: Lowercase file extension or MIME type.
            file_bytes: Raw uploaded bytes.
            file_name: Optional file name used for MIME/extension fallback.
            mime_type: Optional MIME type from upload metadata.

        Returns:
            Tuple containing ordered text segments and the page or sheet count.

        Raises:
            ValueError: If the file type is unsupported.
        """
        extractor_name = self.resolve_extractor_name(
            type_identifier,
            file_name=file_name,
            mime_type=mime_type,
        )
        if extractor_name is None:
            raise ValueError(f"Unsupported file type: {type_identifier}")

        resolved_extension = (
            self.normalize_extension(type_identifier)
            if type_identifier.startswith(".")
            else self.resolve_extension(file_name, mime_type or type_identifier)
        )
        resolved_mime_type = self.resolve_mime_type(
            file_name or f"document{resolved_extension}",
            mime_type or (type_identifier if "/" in type_identifier else None),
        )

        if extractor_name == "read_pdf":
            return self.extract_pdf_segments(file_bytes)
        if extractor_name == "read_excel":
            return self.extract_excel_segments(file_bytes, source_extension=resolved_extension)

        extractor = self.extractors[extractor_name]
        text, page_count = extractor(
            file_bytes,
            file_name=file_name,
            mime_type=resolved_mime_type,
            source_extension=resolved_extension,
        )
        return [text] if text else [""], page_count

    def read_text(self, file_bytes: bytes, **_: object) -> tuple[str, int]:
        """Extract text from a plain-text document."""
        text = self._normalize_text(self._decode_text_bytes(file_bytes))
        return text, 1

    def read_html(self, file_bytes: bytes, **_: object) -> tuple[str, int]:
        """Extract readable text from an HTML document."""
        html_text = self._decode_text_bytes(file_bytes)
        return self._html_to_text(html_text), 1

    def extract_text_from_pdf(self, file_bytes: bytes) -> tuple[str, int]:
        """Extract text and page count from a PDF using PyMuPDF.

        Args:
            file_bytes: Raw PDF bytes.

        Returns:
            Tuple containing extracted text and page count.
        """
        pages, page_count = self.extract_pdf_segments(file_bytes)
        return "\n\n".join(pages), page_count

    def read_pdf(self, file_bytes: bytes, **_: object) -> tuple[str, int]:
        """Extract text from a PDF document."""
        return self.extract_text_from_pdf(file_bytes)

    def extract_pdf_segments(self, file_bytes: bytes) -> tuple[list[str], int]:
        """Extract one text segment per PDF page.

        Args:
            file_bytes: Raw PDF bytes.

        Returns:
            Tuple containing per-page extracted text and the page count.
        """
        import fitz

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        pages: list[str] = []
        for page in doc:
            pages.append(self._normalize_text(page.get_text()))
        doc.close()
        return pages, len(pages)

    def extract_text_from_docx(self, file_bytes: bytes) -> tuple[str, int]:
        """Extract text and an approximate page count from a DOCX file.

        Args:
            file_bytes: Raw DOCX bytes.

        Returns:
            Tuple containing extracted text and estimated page count.
        """
        from docx import Document

        doc = Document(io.BytesIO(file_bytes))
        paragraphs = [paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()]
        text = self._normalize_text("\n\n".join(paragraphs))
        page_estimate = max(1, len(text) // 3000) if text else 1
        return text, page_estimate

    def read_docx(self, file_bytes: bytes, **_: object) -> tuple[str, int]:
        """Extract text from a DOCX document."""
        return self.extract_text_from_docx(file_bytes)

    def extract_excel_segments(
        self,
        file_bytes: bytes,
        *,
        source_extension: str = ".xlsx",
    ) -> tuple[list[str], int]:
        """Extract one text segment per spreadsheet sheet.

        Args:
            file_bytes: Raw Excel bytes.
            source_extension: Canonical spreadsheet extension.

        Returns:
            Tuple containing per-sheet text and the sheet count.
        """
        normalized_extension = self.normalize_extension(source_extension)
        parts: list[str] = []

        if normalized_extension == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_bytes)
            for sheet in workbook.sheets():
                rows: list[str] = []
                for row_idx in range(sheet.nrows):
                    values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                    row_text = self._stringify_row_cells(values)
                    if row_text.strip():
                        rows.append(row_text)
                parts.append(f"--- Sheet: {sheet.name} ---\n" + "\n".join(rows) if rows else "")
            return parts, workbook.nsheets

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows: list[str] = []
                for row in worksheet.iter_rows(values_only=True):
                    row_text = self._stringify_row_cells(list(row))
                    if row_text.strip():
                        rows.append(row_text)
                parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows) if rows else "")
            return parts, len(workbook.sheetnames)
        finally:
            workbook.close()

    def extract_text_from_xlsx(self, file_bytes: bytes) -> tuple[str, int]:
        """Extract text from all sheets in an XLSX workbook."""
        segments, sheet_count = self.extract_excel_segments(file_bytes, source_extension=".xlsx")
        return "\n\n".join(segment for segment in segments if segment), sheet_count

    def extract_xlsx_segments(self, file_bytes: bytes) -> tuple[list[str], int]:
        """Extract one text segment per XLSX sheet."""
        return self.extract_excel_segments(file_bytes, source_extension=".xlsx")

    def read_excel(
        self,
        file_bytes: bytes,
        *,
        source_extension: str | None = None,
        **_: object,
    ) -> tuple[str, int]:
        """Extract text from an Excel workbook."""
        segments, sheet_count = self.extract_excel_segments(
            file_bytes,
            source_extension=source_extension or ".xlsx",
        )
        return "\n\n".join(segment for segment in segments if segment), sheet_count

    def _build_email_text(
        self,
        *,
        subject: str | None,
        sender: str | None,
        to: str | None,
        cc: str | None,
        sent_at: str | None,
        body: str | None,
    ) -> str:
        """Render email metadata and body into readable plain text."""
        parts: list[str] = []
        if subject:
            parts.append(f"Subject: {subject}")
        if sender:
            parts.append(f"From: {sender}")
        if to:
            parts.append(f"To: {to}")
        if cc:
            parts.append(f"Cc: {cc}")
        if sent_at:
            parts.append(f"Date: {sent_at}")
        if body:
            parts.append("")
            parts.append(body)
        return self._normalize_text("\n".join(parts))

    def _extract_email_body(self, message: Message) -> str:
        """Extract the preferred readable body from a parsed RFC822 message."""
        plain_parts: list[str] = []
        html_parts: list[str] = []

        for part in message.walk():
            if part.is_multipart():
                continue

            content_disposition = (part.get_content_disposition() or "").lower()
            if content_disposition == "attachment":
                continue

            content_type = part.get_content_type().lower()
            payload = part.get_payload(decode=True) or b""
            if not payload:
                continue

            charset = part.get_content_charset() or "utf-8"
            try:
                part_text = payload.decode(charset)
            except (LookupError, UnicodeDecodeError):
                part_text = self._decode_text_bytes(payload)

            if content_type == "text/plain":
                plain_parts.append(self._normalize_text(part_text))
            elif content_type == "text/html":
                html_parts.append(self._html_to_text(part_text))

        if plain_parts:
            return self._normalize_text("\n\n".join(filter(None, plain_parts)))
        if html_parts:
            return self._normalize_text("\n\n".join(filter(None, html_parts)))
        return ""

    def read_rfc822(self, file_bytes: bytes, **_: object) -> tuple[str, int]:
        """Extract readable text from an RFC822 `.eml` message."""
        message: EmailMessage = BytesParser(policy=policy.default).parsebytes(file_bytes)
        text = self._build_email_text(
            subject=message.get("subject"),
            sender=message.get("from"),
            to=message.get("to"),
            cc=message.get("cc"),
            sent_at=message.get("date"),
            body=self._extract_email_body(message),
        )
        return text, 1

    def read_outlook(
        self,
        file_bytes: bytes,
        *,
        file_name: str = "",
        **_: object,
    ) -> tuple[str, int]:
        """Extract readable text from an Outlook `.msg` message."""
        import extract_msg

        msg_path = ""
        msg = None
        try:
            # `extract_msg` expects a filesystem path, so stage the bytes briefly.
            with tempfile.NamedTemporaryFile(delete=False, suffix=".msg") as temp_file:
                temp_file.write(file_bytes)
                msg_path = temp_file.name

            msg = extract_msg.Message(msg_path)
            body = msg.body or ""
            if not body and getattr(msg, "htmlBody", None):
                html_body = msg.htmlBody
                if isinstance(html_body, bytes):
                    html_body = self._decode_text_bytes(html_body)
                body = self._html_to_text(str(html_body))

            text = self._build_email_text(
                subject=msg.subject,
                sender=msg.sender,
                to=msg.to,
                cc=msg.cc,
                sent_at=str(msg.date) if msg.date else None,
                body=body,
            )
            return text, 1
        except Exception as exc:
            fallback_name = file_name or "message.msg"
            fallback_text = self._normalize_text(
                "\n".join(
                    [
                        f"Outlook message: {fallback_name}",
                        f"Extractor could not parse the message body: {exc}",
                    ]
                )
            )
            return fallback_text, 1
        finally:
            if msg is not None:
                try:
                    msg.close()
                except Exception:
                    pass
            if msg_path and os.path.exists(msg_path):
                os.unlink(msg_path)

    def _describe_image_dimensions(self, file_bytes: bytes) -> str | None:
        """Return `WIDTHxHEIGHT` when Pillow is available and the image decodes."""
        try:
            from PIL import Image

            with Image.open(io.BytesIO(file_bytes)) as image:
                return f"{image.width}x{image.height}"
        except Exception:
            return None

    def read_image(
        self,
        file_bytes: bytes,
        *,
        file_name: str = "",
        mime_type: str | None = None,
        **_: object,
    ) -> tuple[str, int]:
        """Accept an image upload without OCR and return structured placeholder text."""
        lines = ["[Image document]", "OCR extraction is not enabled for image uploads."]
        if file_name:
            lines.append(f"File name: {file_name}")
        if mime_type:
            lines.append(f"MIME type: {mime_type}")
        dimensions = self._describe_image_dimensions(file_bytes)
        if dimensions:
            lines.append(f"Dimensions: {dimensions}")
        return self._normalize_text("\n".join(lines)), 1

    def extract(self, extension: str, file_bytes: bytes) -> tuple[str, int]:
        """Extract text using the registered extractor for a file type.

        Args:
            extension: Lowercase file extension including the leading dot.
            file_bytes: Raw uploaded bytes.

        Returns:
            Tuple containing extracted text and page or sheet count.

        Raises:
            ValueError: If the extension is unsupported.
        """
        segments, page_count = self.extract_segments(extension, file_bytes)
        return "\n\n".join(segment for segment in segments if segment), page_count

    def count_tokens(self, text: str) -> int:
        """Count the number of tokens in a text string using tiktoken cl100k_base.

        Uses a class-level cached encoder so the encoding is only loaded once
        across all calls and instances.

        Args:
            text: The text to tokenize.

        Returns:
            Number of tokens in the text.

        Example:
            svc = TextExtractionService()
            svc.count_tokens("Hello, world!")
        """
        if not text:
            return 0

        # Lazy-load the encoding once and cache on the class.
        if TextExtractionService._tiktoken_encoding is None:
            import tiktoken

            TextExtractionService._tiktoken_encoding = tiktoken.get_encoding("cl100k_base")

        return len(TextExtractionService._tiktoken_encoding.encode(text))

    def tag_segments(
        self,
        segments: list[str],
        file_name: str,
        content_id: str,
    ) -> list[str]:
        """Prepend a structured metadata tag to each text segment.

        Inserts a ``[DOC_META ...]`` line at the top of every non-empty
        segment so downstream LLMs can identify the source document and
        page or sheet number without relying on positional heuristics.

        Args:
            segments: Ordered text segments (one per page, sheet, or message).
            file_name: Human-readable document file name.
            content_id: Stable backend identifier for the document.

        Returns:
            New list of segments with metadata tags prepended.

        Example:
            svc = TextExtractionService()
            svc.tag_segments(
                ["Page one text", "Page two text"],
                file_name="policy.pdf",
                content_id="abc123",
            )
        """
        tagged: list[str] = []
        for idx, segment in enumerate(segments, start=1):
            if not segment:
                tagged.append(segment)
                continue
            tag = f'[DOC_META content_id="{content_id}" page={idx} doc_name="{file_name}"]'
            tagged.append(f"{tag}\n{segment}")
        return tagged

    def format_file_size(self, size_bytes: int) -> str:
        """Return a human-readable file size string.

        Args:
            size_bytes: Raw file size in bytes.

        Returns:
            Formatted kilobyte or megabyte string.
        """
        if size_bytes > 1024 * 1024:
            return f"{size_bytes / 1024 / 1024:.1f} MB"
        return f"{size_bytes / 1024:.1f} KB"

    def build_document_payload(
        self,
        file_name: str,
        file_bytes: bytes,
        content_id: str | None = None,
        content_url: str | None = None,
        mime_type: str | None = None,
    ) -> dict[str, object]:
        """Build a frontend-facing document payload for one file.

        Each page or sheet segment is prefixed with a ``[DOC_META ...]`` tag so
        downstream LLMs can identify the source document and page number.

        Args:
            file_name: File name on disk.
            file_bytes: File content bytes.
            content_id: Optional stable backend document identifier.
            content_url: Optional public URL for preview/download access.
            mime_type: Optional MIME type supplied by the caller.

        Returns:
            JSON-serializable payload matching the existing frontend contract.
        """
        resolved_extension = self.resolve_extension(file_name, mime_type)
        safe_id = content_id or ""
        segments, page_count = self.extract_segments(
            resolved_extension,
            file_bytes,
            file_name=file_name,
            mime_type=mime_type,
        )
        tagged = self.tag_segments(segments, file_name, safe_id)
        text = "\n\n".join(s for s in tagged if s)
        size = len(file_bytes)
        resolved_mime_type = self.resolve_mime_type(file_name, mime_type)
        return {
            "file_name": file_name,
            "mime_type": resolved_mime_type,
            "content": text,
            "page_count": page_count,
            "file_size": self.format_file_size(size),
            "file_size_bytes": size,
            "token_count": self.count_tokens(text),
            "content_id": safe_id,
            "content_url": content_url or "",
            "path": content_url or "",
        }

    def build_example_documents_from_directory(self, upload_dir: str) -> list[dict[str, object]]:
        """Build example-document payloads for all supported files in a directory.

        Args:
            upload_dir: Directory containing the example source files.

        Returns:
            List of JSON-serializable payloads matching the frontend contract.
        """
        documents: list[dict[str, object]] = []
        try:
            for file_name in sorted(os.listdir(upload_dir)):
                extension = self.get_extension(file_name)
                if extension not in self.allowed_extensions:
                    continue

                file_path = os.path.join(upload_dir, file_name)
                if not os.path.isfile(file_path):
                    continue

                with open(file_path, "rb") as file_obj:
                    file_bytes = file_obj.read()

                documents.append(self.build_document_payload(file_name, file_bytes))
        except Exception as exc:
            print(f"[EXAMPLE-DOCS] Error scanning uploads dir: {exc}", flush=True)

        return documents

    def build_upload_response(
        self,
        file_name: str,
        file_bytes: bytes,
        content_id: str | None = None,
        content_url: str | None = None,
        mime_type: str | None = None,
    ) -> dict[str, object]:
        """Build the upload response payload after extraction.

        Each page or sheet segment is prefixed with a ``[DOC_META ...]`` tag so
        downstream LLMs can identify the source document and page number.

        Args:
            file_name: Uploaded file name.
            file_bytes: Raw uploaded bytes.
            content_id: Optional stable backend document identifier.
            content_url: Optional public URL for preview/download access.
            mime_type: Optional MIME type supplied by the upload client.

        Returns:
            JSON-serializable payload for ``POST /upload``.
        """
        resolved_extension = self.resolve_extension(file_name, mime_type)
        safe_id = content_id or ""
        segments, page_count = self.extract_segments(
            resolved_extension,
            file_bytes,
            file_name=file_name,
            mime_type=mime_type,
        )
        tagged = self.tag_segments(segments, file_name, safe_id)
        extracted_text = "\n\n".join(s for s in tagged if s)
        file_size = len(file_bytes)
        resolved_mime_type = self.resolve_mime_type(file_name, mime_type)
        return {
            "filename": file_name,
            "file_type": resolved_extension.lstrip("."),
            "mime_type": resolved_mime_type,
            "file_size": self.format_file_size(file_size),
            "file_size_bytes": file_size,
            "page_count": page_count,
            "content": extracted_text,
            "token_count": self.count_tokens(extracted_text),
            "content_id": safe_id,
            "content_url": content_url or "",
            "path": content_url or "",
        }
