"""
Audit Assistant Agent - AG-UI endpoint using Pydantic AI.

This creates an AG-UI compatible endpoint that CopilotKit can connect to
for bidirectional state synchronization and streaming agent responses.
Also provides a file upload endpoint for document processing.

uv run uvicorn main:app --reload --port 8001
"""

import os
import re
import json
from dotenv import load_dotenv

# Load environment variables first
load_dotenv()

from starlette.routing import Route  # noqa: E402
from starlette.responses import JSONResponse, StreamingResponse  # noqa: E402
from starlette.applications import Starlette  # noqa: E402
from starlette.requests import Request  # noqa: E402

from agent import agent, AuditState  # noqa: E402
from pydantic_ai.ag_ui import StateDeps  # noqa: E402

# Configuration
BACKEND_PORT = int(os.getenv("BACKEND_PORT", "8001"))
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def pascal_to_screaming_snake(name: str) -> str:
    """Convert PascalCase to SCREAMING_SNAKE_CASE.

    Args:
        name: PascalCase event type name.

    Returns:
        SCREAMING_SNAKE_CASE version of the name.
    """
    result = re.sub(r'(?<!^)(?=[A-Z])', '_', name)
    return result.upper()


# Map of PascalCase to SCREAMING_SNAKE_CASE event types
EVENT_TYPE_MAP = {
    "RunStarted": "RUN_STARTED",
    "RunFinished": "RUN_FINISHED",
    "RunError": "RUN_ERROR",
    "StepStarted": "STEP_STARTED",
    "StepFinished": "STEP_FINISHED",
    "StateSnapshot": "STATE_SNAPSHOT",
    "StateDelta": "STATE_DELTA",
    "MessagesSnapshot": "MESSAGES_SNAPSHOT",
    "TextMessageStart": "TEXT_MESSAGE_START",
    "TextMessageContent": "TEXT_MESSAGE_CONTENT",
    "TextMessageEnd": "TEXT_MESSAGE_END",
    "TextMessageChunk": "TEXT_MESSAGE_CHUNK",
    "ToolCallStart": "TOOL_CALL_START",
    "ToolCallArgs": "TOOL_CALL_ARGS",
    "ToolCallEnd": "TOOL_CALL_END",
    "ToolCallChunk": "TOOL_CALL_CHUNK",
    "ToolCallResult": "TOOL_CALL_RESULT",
    "Raw": "RAW",
    "Custom": "CUSTOM",
}


def transform_event_type(event_data: str) -> str:
    """Transform event type in SSE data from PascalCase to SCREAMING_SNAKE_CASE.

    Args:
        event_data: JSON string containing an event with a 'type' field.

    Returns:
        JSON string with the 'type' field converted to SCREAMING_SNAKE_CASE.
    """
    try:
        data = json.loads(event_data)
        if "type" in data:
            original_type = data["type"]
            if original_type in EVENT_TYPE_MAP:
                data["type"] = EVENT_TYPE_MAP[original_type]
            elif not original_type.isupper():
                data["type"] = pascal_to_screaming_snake(original_type)
        return json.dumps(data)
    except json.JSONDecodeError:
        return event_data


async def transform_sse_stream(original_response):
    """Transform SSE stream to use SCREAMING_SNAKE_CASE event types.

    Wraps the original Pydantic AI SSE response and converts event names
    and data JSON 'type' fields for CopilotKit v2 compatibility.

    Args:
        original_response: The original StreamingResponse from pydantic-ai.

    Yields:
        Transformed SSE chunks with SCREAMING_SNAKE_CASE event types.
    """
    try:
        async for chunk in original_response.body_iterator:
            try:
                if isinstance(chunk, bytes):
                    chunk = chunk.decode('utf-8')

                lines = chunk.split('\n')
                transformed_lines = []

                for line in lines:
                    if line.startswith('event: '):
                        event_name = line[7:]
                        if event_name in EVENT_TYPE_MAP:
                            transformed_lines.append(f'event: {EVENT_TYPE_MAP[event_name]}')
                        elif not event_name.isupper():
                            transformed_lines.append(
                                f'event: {pascal_to_screaming_snake(event_name)}'
                            )
                        else:
                            transformed_lines.append(line)
                    elif line.startswith('data: '):
                        data = line[6:]
                        transformed_data = transform_event_type(data)
                        transformed_lines.append(f'data: {transformed_data}')
                    else:
                        transformed_lines.append(line)

                yield '\n'.join(transformed_lines)
            except Exception as chunk_error:
                print(f"[SSE ERROR] Error processing chunk: {chunk_error}", flush=True)
                import traceback
                traceback.print_exc()
                raise
    except Exception as stream_error:
        print(f"[SSE ERROR] Stream error: {stream_error}", flush=True)
        import traceback
        traceback.print_exc()
        raise


# Create the base AG-UI app using Pydantic AI's built-in integration
_base_ag_ui_app = agent.to_ag_ui(
    deps=StateDeps(AuditState()),
)

# Create our wrapper app
app = Starlette()


# AG-UI POST endpoint with event type transformation
async def ag_ui_endpoint(request: Request):
    """AG-UI endpoint that wraps Pydantic AI's implementation
    and transforms event types to SCREAMING_SNAKE_CASE format.
    """
    try:
        body = await request.body()
        print(f"[AG-UI] Received request: {len(body)} bytes", flush=True)

        from starlette.requests import Request as StarletteRequest

        scope = dict(request.scope)

        async def receive():
            return {"type": "http.request", "body": body}

        new_request = StarletteRequest(scope, receive)

        # Find the POST route handler in the base AG-UI app
        for route in _base_ag_ui_app.routes:
            if hasattr(route, 'methods') and 'POST' in route.methods:
                original_response = await route.endpoint(new_request)
                break
        else:
            return JSONResponse({"error": "AG-UI endpoint not found"}, status_code=500)

        # Wrap streaming responses with our event type transformer
        if isinstance(original_response, StreamingResponse):
            return StreamingResponse(
                transform_sse_stream(original_response),
                media_type="text/event-stream",
                headers=dict(original_response.headers),
            )

        return original_response
    except Exception as e:
        print(f"[AG-UI ERROR] {e}", flush=True)
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


# =========================================================================
# Text extraction helpers
# =========================================================================

def extract_text_from_pdf(file_bytes: bytes) -> tuple[str, int]:
    """Extract text and page count from a PDF using PyMuPDF.

    Args:
        file_bytes: Raw PDF bytes.

    Returns:
        Tuple of (extracted_text, page_count).
    """
    import fitz  # PyMuPDF

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages: list[str] = []
    for page in doc:
        pages.append(page.get_text())
    doc.close()
    return "\n\n".join(pages), len(pages)


def extract_text_from_docx(file_bytes: bytes) -> tuple[str, int]:
    """Extract text and approximate page count from a DOCX file.

    Args:
        file_bytes: Raw DOCX bytes.

    Returns:
        Tuple of (extracted_text, estimated_page_count).
    """
    import io
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n\n".join(paragraphs)
    # Rough page estimate: ~3000 chars per page
    page_estimate = max(1, len(text) // 3000)
    return text, page_estimate


def extract_text_from_xlsx(file_bytes: bytes) -> tuple[str, int]:
    """Extract text from an XLSX file by reading all sheets.

    Each row is joined with tabs, each sheet separated by a header line.

    Args:
        file_bytes: Raw XLSX bytes.

    Returns:
        Tuple of (extracted_text, sheet_count).
    """
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts), len(wb.sheetnames)


EXTRACTORS = {
    ".pdf": extract_text_from_pdf,
    ".docx": extract_text_from_docx,
    ".xlsx": extract_text_from_xlsx,
}


# =========================================================================
# File upload endpoint
# =========================================================================

async def upload_endpoint(request: Request):
    """Handle document file uploads with text extraction.

    Accepts multipart form data with a 'file' field. Supports
    .pdf, .docx, and .xlsx file types. Extracts text content so
    the agent can analyze the document.

    Returns:
        JSON with file metadata *and* extracted text content.
    """
    try:
        form = await request.form()
        file = form.get("file")

        if not file:
            return JSONResponse({"error": "No file provided"}, status_code=400)

        # Validate file type
        allowed_extensions = {".pdf", ".docx", ".xlsx"}
        filename = file.filename or "unknown"
        ext = os.path.splitext(filename)[1].lower()

        if ext not in allowed_extensions:
            return JSONResponse(
                {"error": f"Unsupported file type: {ext}. Allowed: {', '.join(allowed_extensions)}"},
                status_code=400,
            )

        # Read file bytes
        file_bytes = await file.read()

        # Save to disk (useful for debugging / future re-processing)
        file_path = os.path.join(UPLOAD_DIR, filename)
        with open(file_path, "wb") as f:
            f.write(file_bytes)

        # Extract text content
        extractor = EXTRACTORS.get(ext)
        if extractor:
            extracted_text, page_count = extractor(file_bytes)
        else:
            extracted_text, page_count = "", 0

        print(
            f"[UPLOAD] {filename}: {len(file_bytes)} bytes, "
            f"{page_count} pages, {len(extracted_text)} chars extracted",
            flush=True,
        )

        # Human-readable file size
        file_size = len(file_bytes)
        size_str = (
            f"{file_size / 1024 / 1024:.1f} MB"
            if file_size > 1024 * 1024
            else f"{file_size / 1024:.1f} KB"
        )

        return JSONResponse({
            "filename": filename,
            "file_type": ext.lstrip("."),
            "file_size": size_str,
            "file_size_bytes": file_size,
            "page_count": page_count,
            "content": extracted_text,
            "path": file_path,
        })
    except Exception as e:
        print(f"[UPLOAD ERROR] {e}", flush=True)
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


# Info endpoint for debugging / CopilotKit compatibility
async def info_endpoint(request: Request):
    """Return agent information."""
    return JSONResponse({
        "name": "Audit Assistant Agent",
        "version": "1.0.0",
        "protocol": "ag-ui",
        "description": "Analyze documents and generate custom audit questionnaires",
    })


# Health endpoint
async def health_endpoint(request: Request):
    """Health check endpoint."""
    return JSONResponse({
        "status": "healthy",
        "agent_ready": bool(os.getenv("OPENAI_API_KEY")),
    })


# GET handler for root
async def root_get_endpoint(request: Request):
    """Return AG-UI endpoint info for GET requests."""
    return JSONResponse({
        "protocol": "ag-ui",
        "version": "1.0.0",
        "endpoints": {
            "run_agent": "POST /",
            "upload": "POST /upload",
            "info": "GET /info",
            "health": "GET /health",
        },
        "description": "Audit Assistant Agent - POST to / to run the agent",
    })


# Combined handler for root path
async def root_handler(request: Request):
    """Handle both GET and POST for root path."""
    if request.method == "POST":
        return await ag_ui_endpoint(request)
    return await root_get_endpoint(request)


# Add routes to the app
app.routes.append(Route("/", root_handler, methods=["GET", "POST"]))
app.routes.append(Route("/upload", upload_endpoint, methods=["POST"]))
app.routes.append(Route("/info", info_endpoint, methods=["GET"]))
app.routes.append(Route("/health", health_endpoint, methods=["GET"]))


# Add CORS middleware for development
from starlette.middleware.cors import CORSMiddleware  # noqa: E402

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    """Startup event handler."""
    print(f"[*] Audit Assistant Agent (AG-UI) starting on port {BACKEND_PORT}")
    print(f"[*] AG-UI endpoint: POST http://localhost:{BACKEND_PORT}/")
    print(f"[*] Upload endpoint: POST http://localhost:{BACKEND_PORT}/upload")
    print(f"[*] Info endpoint: GET http://localhost:{BACKEND_PORT}/info")
    print(f"[*] Health endpoint: GET http://localhost:{BACKEND_PORT}/health")

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("[!] WARNING: OPENAI_API_KEY not set")
    else:
        print("[+] OpenAI API key configured")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=BACKEND_PORT, reload=True)
